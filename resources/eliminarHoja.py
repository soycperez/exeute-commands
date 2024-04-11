import sys
from openpyxl import load_workbook 
import pandas as pd 

def eliminarHoja(pathExcel, nombrehoja):
    print('Ruta excel:', pathExcel)
    print('Hoja a eliminar', nombrehoja)

    dataframe = pd.read_excel(pathExcel)
    print('Excel recuperado')
    with pd.ExcelWriter(pathExcel, engine='openpyxl', mode='a') as writer: 
        workBook = writer.book
        print(workBook.sheetnames)
        try: 
            workBook.remove(workBook[nombrehoja])
            print("OK")
            print('Se remueve la hoja ', nombrehoja)
        except:
            print('ERROR')
            print('No existe la hoja en este archivo')
        finally:
            print(workBook.sheetnames)
            #dataframe.to_excel(writer)

if len(sys.argv)<=1:
    print('ERROR')
    print('Indique parametros')
else:
    eliminarHoja(sys.argv[1], sys.argv[2])
        